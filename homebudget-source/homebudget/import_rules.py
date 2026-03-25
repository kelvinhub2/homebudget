#!/usr/bin/env python3
"""
import_rules.py — One-time import of rules and taxonomy from XLSX files
Run on Pi5 INSIDE the container:

  docker exec homebudget python3 /app/import_rules.py \
    --rules /nextcloud/docker-configs/homebudget/import_rules.py \
    --taxonomy /nextcloud/docker-configs/homebudget/taxonomy.yaml

Or copy files to /tmp first, then run:
  docker cp import_rules.py homebudget:/app/
  docker cp rules_0_1.xlsx homebudget:/tmp/
  docker cp taxonomy.yaml homebudget:/tmp/
  docker exec homebudget python3 /app/import_rules.py

Options:
  --rules-xlsx  PATH   Path to rules_0_1.xlsx (default: /tmp/rules_0_1.xlsx)
  --taxonomy    PATH   Path to taxonomy.yaml  (default: /tmp/taxonomy.yaml)
  --dry-run            Print what would be imported, don't write to DB
  --clear-rules        Delete all existing rules before import
  --clear-taxonomy     Delete all existing taxonomy before import
"""

import sys
import os
import re
import argparse
import yaml

# Add app directory to path so we can import db.py
sys.path.insert(0, "/app")
import db

# ── L1 name mapping: old → new taxonomy ──────────────────────────────────────
L1_MAP = {
    # rules_0_1.xlsx names → our taxonomy
    "Finance & Admin":   "Finance & Admin",
    "Charity":           "Leisure",           # Donations → Leisure/Donations
    "Family":            "Children",          # Nanny rules → Children
    "Holiday Home":      "Home",              # → Home/Holiday Home
    "Personal Care":     "Health",            # → Health/Personal Care
    "Savings":           "Finance & Admin",   # Pension → Finance & Admin
    # Unchanged
    "Income":            "Income",
    "Daily Living":      "Daily Living",
    "Home":              "Home",
    "Health":            "Health",
    "Children":          "Children",
    "Transportation":    "Transportation",
    "Leisure":           "Leisure",
    "Shopping":          "Leisure",           # → Leisure/Shopping & Gadgets
    # rules.xlsx (v1) names
    "Home Expenses":     "Home",
    "Transfers":         "Finance & Admin",
    "Tax":               "Finance & Admin",
}

# ── L2 name mapping: old → new ───────────────────────────────────────────────
L2_MAP = {
    # Consolidations
    "Bakery":                "Groceries",
    "Butcher":               "Groceries",
    "Lunch":                 "Canteen",
    "Grooming":              "Personal Care",
    "Hairdresser":           "Personal Care",
    "Children's Health":     "Doctors Kids",
    "Diagnostics":           "Doctors",
    "GP":                    "Doctors",
    "Specialist":            "Doctors",
    "Children / Nanny":      "Nanny",
    "Nanny / Childcare":     "Nanny",
    "Shared Expenses":       "Shared Expenses",
    "Music Lessons":         "Activities",     # under Children
    "Car Maintenance":       "Car Repair",
    "Car Inspection":        "Car Repair",
    "Car Train":             "Public Transport",
    "Airport":               "Other",
    "Car":                   "Car Repair",
    # Leisure consolidations
    "Going Out":             "Dining Out",
    "Restaurant":            "Dining Out",
    "Food":                  "Dining Out",
    "Accommodation":         "Hotel",
    "Flights":               "Travel",
    "Holiday Transportation":"Travel",
    "Travel Japan":          "Travel",
    "Abroad":                "Shopping & Gadgets",
    "Culture":               "Entertainment",
    "Ice Skating":           "Sport & Activities",
    "Skiing":                "Sport & Activities",
    "Outdoor":               "Sport & Activities",
    "Hiking":                "Sport & Activities",
    "Sport":                 "Sport & Activities",
    "Sport & Fitness":       "Sport & Activities",
    "Gaming":                "Entertainment",
    "Community":             "Entertainment",
    "Tickets":               "Entertainment",
    "Events":                "Entertainment",
    "Event":                 "Entertainment",
    "Electronics":           "Shopping & Gadgets",
    "Gadgets":               "Shopping & Gadgets",
    "Shopping":              "Shopping & Gadgets",
    "Gifts":                 "Shopping & Gadgets",
    "Books":                 "Shopping & Gadgets",
    "Books/School":          "Shopping & Gadgets",
    "Clothing":              "Shopping & Gadgets",
    "Clothing & Shoes":      "Shopping & Gadgets",
    "Flowers":               "Shopping & Gadgets",
    "Jewelry & Watches":     "Shopping & Gadgets",
    "Local":                 "Shopping & Gadgets",
    "Marketplace":           "Shopping & Gadgets",
    "Online":                "Shopping & Gadgets",
    "Online Shopping":       "Shopping & Gadgets",
    "Software":              "Shopping & Gadgets",
    "Office Supplies":       "Shopping & Gadgets",
    "Spain":                 "Shopping & Gadgets",
    "Sport (shopping)":      "Shopping & Gadgets",
    "Subscriptions (shopping)": "Subscriptions",
    # Finance & Admin
    "Pension 2nd Pillar":    "Pension",
    "Pension 3rd Pillar":    "Pension",
    "Social Charges Nanny":  "Other",
    "TWINT Top-up":          "Other",
    "TWINT Voucher":         "Other",
    "Work Subscriptions":    "Other",
    "Telecom":               "Other",
    "Wealth Mgmt":           "Investments",
    "Municipal Fees":        "Tax",
    "Official Fees":         "Tax",
    "Customs & Duties":      "Customs & Duties",
    "Cash Withdrawal":       "ATM",
    "Banking":               "Banking Fees",
    "Bank Interest":         "Other",        # under Income
    # Home
    "Cleaning (holiday)":    "Holiday Home",
    "Legal":                 "Holiday Home",
    "Expenses":              "Holiday Home",
    "Rent (holiday)":        "Holiday Home",
    "DIY & Garden":          "Garden",
    "Utilities":             "Electricity",
    "Bank Fees":             "Banking Fees",
    "Medical":               "Doctors",
    "Children":              "Shopping & Gadgets",   # Shopping/Children
    "Membership":            "Other",
    "Networking":            "Smart Home",
    "Services":              "Other",
    "Interior Design":       "Furniture",
    "Renovation":            "Repairs",
    # Holiday Home L2 → Home/Holiday Home
    "ATM (holiday)":         "Holiday Home",
    # Donations
    "Donations":             "Donations",
    # Keep as-is
    "Salary":                "Salary",
    "Dividends":             "Dividends",
    "Household Refunds":     "Household Refunds",
    "Rental Income":         "Rental Income",
    "Solar":                 "Solar",
    "Marketplace Sale":      "Marketplace Sale",
    "Expense Reimbursement": "Expense Reimbursement",
    "Bank Interest":         "Bank Interest",
    "Groceries":             "Groceries",
    "Canteen":               "Canteen",
    "Dining Out":            "Dining Out",
    "Food Delivery":         "Food Delivery",
    "Shared Expenses":       "Shared Expenses",
    "Cash Out":              "Cash Out",
    "Mortgage":              "Mortgage",
    "Electricity":           "Electricity",
    "Gas":                   "Gas",
    "Water":                 "Water",
    "TV & Internet":         "TV & Internet",
    "Internet & TV":         "TV & Internet",
    "TV Internet":           "TV & Internet",
    "Furniture":             "Furniture",
    "Garden":                "Garden",
    "Smart Home":            "Smart Home",
    "Insurance":             "Insurance",
    "Home Insurance":        "Insurance",
    "Health Insurance":      "Health Insurance",
    "Doctors":               "Doctors",
    "Doctors Kids":          "Doctors Kids",
    "Pharmacy":              "Pharmacy",
    "Drugs":                 "Pharmacy",
    "Lab":                   "Pharmacy",
    "Optician":              "Optician",
    "Wellness":              "Wellness",
    "Personal Care":         "Personal Care",
    "Nanny":                 "Nanny",
    "Childcare":             "Childcare",
    "Activities":            "Activities",
    "Education":             "Education",
    "Fuel":                  "Fuel",
    "Car Repair":            "Car Repair",
    "Car Insurance":         "Car Insurance",
    "Car Tax":               "Car Tax",
    "Public Transport":      "Public Transport",
    "Parking":               "Parking",
    "Taxi":                  "Taxi",
    "Travel":                "Travel",
    "Hotel":                 "Hotel",
    "Subscriptions":         "Subscriptions",
    "Entertainment":         "Entertainment",
    "Sport & Activities":    "Sport & Activities",
    "Shopping & Gadgets":    "Shopping & Gadgets",
    "P2P Transfer":          "P2P Transfer",
    "Donations":             "Donations",
    "Tax":                   "Tax",
    "Fines":                 "Fines",
    "Banking Fees":          "Banking Fees",
    "CC Settlement":         "CC Settlement",
    "ATM":                   "ATM",
    "Internal Transfer":     "Internal Transfer",
    "Pension":               "Pension",
    "Investments":           "Investments",
    "Customs & Duties":      "Customs & Duties",
    "Unknown":               "Unknown",
}


def map_l1(l1: str) -> str:
    return L1_MAP.get(l1, l1)


def map_l2(l2: str, new_l1: str) -> str:
    """Map old L2 to new L2, with L1 context for ambiguous cases."""
    if not l2:
        return "Other"

    # Holiday Home sub-categories → Home/Holiday Home
    if new_l1 == "Home" and l2 in ("ATM", "Cleaning", "Expenses", "Legal", "Rent"):
        return "Holiday Home"

    # Charity/Donations → Leisure/Donations
    if l2 == "Donations":
        return "Donations"

    # Family Shared Expenses → Daily Living
    if l2 == "Shared Expenses":
        return "Shared Expenses"

    mapped = L2_MAP.get(l2)
    if mapped:
        return mapped

    # Unknown L2 — keep as-is but warn
    return l2


def parse_keywords(raw: str) -> tuple[str, str]:
    """
    Split semicolon-separated keywords.
    First part → merchant field (primary match string)
    All parts → stored as keyword alternatives
    Returns (merchant, keyword_string)
    """
    if not raw:
        return ("", "")
    parts = [k.strip() for k in str(raw).split(";") if k.strip()]
    if not parts:
        return ("", "")
    merchant = parts[0]
    # keyword field: all parts joined (for multi-keyword matching)
    keyword = ";".join(parts)
    return merchant, keyword


def load_taxonomy(yaml_path: str) -> dict:
    """Load taxonomy from YAML file."""
    with open(yaml_path, "r", encoding="utf-8") as f:
        data = yaml.safe_load(f)
    return data.get("taxonomy", {})


def import_taxonomy(yaml_path: str, clear: bool, dry_run: bool):
    """Import L1/L2 taxonomy from YAML into DB."""
    taxonomy = load_taxonomy(yaml_path)
    count = 0

    if not dry_run and clear:
        with db.get_conn() as conn:
            conn.execute("DELETE FROM taxonomy")
        print("  Cleared existing taxonomy")

    with db.get_conn() as conn:
        for l1, l1_data in taxonomy.items():
            l2_list = l1_data.get("l2", [])
            for l2 in l2_list:
                if dry_run:
                    print(f"  [DRY] taxonomy: {l1} / {l2}")
                else:
                    conn.execute(
                        "INSERT OR IGNORE INTO taxonomy (l1, l2) VALUES (?, ?)",
                        (l1, l2)
                    )
                count += 1

    print(f"  Taxonomy: {count} L2 entries {'(dry run)' if dry_run else 'imported'}")
    return count


def import_rules(xlsx_path: str, clear: bool, dry_run: bool):
    """Import rules from XLSX into DB."""
    from openpyxl import load_workbook

    wb   = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws   = wb["Rules"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("  ERROR: No rows found in Rules sheet")
        return 0

    # Detect header
    header = [str(h or "").strip().lower() for h in rows[0]]
    print(f"  Columns: {header}")

    def col(names):
        for n in names:
            for i, h in enumerate(header):
                if n.lower() in h:
                    return i
        return None

    idx_kw      = col(["keywords", "keyword"])
    idx_l1      = col(["l1"])
    idx_l2      = col(["l2"])
    idx_merchant= col(["merchant_clean", "merchant"])
    idx_issub   = col(["issub", "is_sub", "sub"])

    if None in (idx_kw, idx_l1, idx_l2):
        print(f"  ERROR: Could not find required columns. Header: {header}")
        return 0

    if not dry_run and clear:
        with db.get_conn() as conn:
            conn.execute("DELETE FROM rules")
        print("  Cleared existing rules")

    imported   = 0
    skipped    = 0
    unmapped   = []

    for row_num, row in enumerate(rows[1:], start=2):
        raw_kw   = row[idx_kw]      if idx_kw      is not None else None
        raw_l1   = row[idx_l1]      if idx_l1      is not None else None
        raw_l2   = row[idx_l2]      if idx_l2      is not None else None
        raw_merch= row[idx_merchant] if idx_merchant is not None else None
        raw_sub  = row[idx_issub]   if idx_issub   is not None else None

        if not raw_kw or not raw_l1:
            skipped += 1
            continue

        # Skip separator/comment rows
        kw_str = str(raw_kw).strip()
        if kw_str.startswith("■") or kw_str.startswith("#"):
            skipped += 1
            continue

        old_l1 = str(raw_l1).strip()
        old_l2 = str(raw_l2).strip() if raw_l2 else "Other"

        new_l1 = map_l1(old_l1)
        new_l2 = map_l2(old_l2, new_l1)

        merchant, keyword = parse_keywords(kw_str)
        is_sub = 1 if str(raw_sub or "").strip().lower() in ("yes", "1", "true") else 0
        merchant_clean = str(raw_merch).strip() if raw_merch else merchant

        # Track unmapped
        if new_l2 == old_l2 and old_l2 not in L2_MAP.values():
            unmapped.append(f"row {row_num}: {old_l1}/{old_l2} → {new_l1}/{new_l2}")

        priority = row_num  # preserve original order = priority

        if dry_run:
            print(f"  [DRY] row {row_num:3d}: {old_l1}/{old_l2} → {new_l1}/{new_l2} | {merchant[:40]}")
        else:
            with db.get_conn() as conn:
                conn.execute("""
                    INSERT OR IGNORE INTO rules
                    (priority, merchant, keyword, l1, l2, is_recurring, active)
                    VALUES (?, ?, ?, ?, ?, ?, 1)
                """, (priority, merchant, keyword, new_l1, new_l2, is_sub))

        imported += 1

    print(f"  Rules: {imported} imported, {skipped} skipped {'(dry run)' if dry_run else ''}")

    if unmapped:
        print(f"\n  WARNING: {len(unmapped)} L2 values not in mapping table:")
        for u in unmapped[:10]:
            print(f"    {u}")
        if len(unmapped) > 10:
            print(f"    ... and {len(unmapped)-10} more")

    return imported


def main():
    parser = argparse.ArgumentParser(description="Import rules and taxonomy into HomeBudget DB")
    parser.add_argument("--rules-xlsx",   default="/tmp/rules_0_1.xlsx", help="Path to rules XLSX")
    parser.add_argument("--taxonomy",     default="/tmp/taxonomy.yaml",   help="Path to taxonomy YAML")
    parser.add_argument("--dry-run",      action="store_true",            help="Preview only, no DB writes")
    parser.add_argument("--clear-rules",  action="store_true",            help="Delete existing rules first")
    parser.add_argument("--clear-taxonomy", action="store_true",          help="Delete existing taxonomy first")
    parser.add_argument("--skip-taxonomy", action="store_true",           help="Skip taxonomy import")
    parser.add_argument("--skip-rules",    action="store_true",           help="Skip rules import")
    args = parser.parse_args()

    print("=== HomeBudget — Rules & Taxonomy Import ===")
    if args.dry_run:
        print("DRY RUN — no changes will be written\n")

    db.init_db()

    if not args.skip_taxonomy:
        if not os.path.exists(args.taxonomy):
            print(f"ERROR: taxonomy file not found: {args.taxonomy}")
            sys.exit(1)
        print(f"\n[1/2] Importing taxonomy from {args.taxonomy}")
        import_taxonomy(args.taxonomy, args.clear_taxonomy, args.dry_run)

    if not args.skip_rules:
        if not os.path.exists(args.rules_xlsx):
            print(f"ERROR: rules file not found: {args.rules_xlsx}")
            sys.exit(1)
        print(f"\n[2/2] Importing rules from {args.rules_xlsx}")
        import_rules(args.rules_xlsx, args.clear_rules, args.dry_run)

    print("\n=== Done ===")
    if not args.dry_run:
        # Show summary
        with db.get_conn() as conn:
            r = conn.execute("SELECT COUNT(*) FROM rules WHERE active=1").fetchone()[0]
            t = conn.execute("SELECT COUNT(*) FROM taxonomy").fetchone()[0]
            print(f"  Active rules:    {r}")
            print(f"  Taxonomy entries:{t}")
            print(f"\nNext: go to Rules page → RE-RUN ALL to re-categorize existing transactions")


if __name__ == "__main__":
    main()
