"""
parsers/swisscard.py — Swisscard CSV + XLSX parser

Supported formats:
  1. CSV (old): TransactionId, CardId, Date, Amount, MerchantName, ...
  2. XLSX (new): Transaction date, Description, Merchant, Card number,
                 Currency, Amount, Foreign Currency, Amount in foreign currency,
                 Debit/Credit, Status, Merchant Category, Registered Category
"""
import re
import csv
import io
from datetime import datetime

from normalizer import extract_merchant_swisscard
from db import make_tx_id


def _merchant_key(name: str) -> str:
    """Extract first alphabetic token for consistent dedup across CSV/XLSX formats."""
    tokens = re.split(r'[^A-Za-z]', name or "")
    first  = next((t for t in tokens if len(t) > 1), tokens[0] if tokens else name)
    return first.upper()[:15]


def is_swisscard_file(filepath: str) -> bool:
    """Detect Swisscard CSV or XLSX by header."""
    try:
        if filepath.lower().endswith(".xlsx"):
            return _is_swisscard_xlsx(filepath)
        with open(filepath, "r", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            return "TransactionId" in headers and "CardId" in headers
    except Exception:
        return False


def _is_swisscard_xlsx(filepath: str) -> bool:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
        wb.close()
        return "Transaction date" in headers and "Merchant" in headers
    except Exception:
        return False


def extract_card_id(filepath: str) -> str:
    """Extract card ID from first data row."""
    try:
        if filepath.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            wb = load_workbook(filepath, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(max_row=3, values_only=True))
            wb.close()
            if len(rows) >= 2:
                # Card number is column index 3
                val = str(rows[1][3] or "")
                # Strip to last 4 digits for account matching
                return re.sub(r'[\s\*]', '', val)
        else:
            with open(filepath, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    return (row.get("CardId") or "").strip()
    except Exception:
        pass
    return ""


def parse(filepath: str, account_id: str, import_id: int) -> list[dict]:
    """Parse Swisscard CSV or XLSX file."""
    if filepath.lower().endswith(".xlsx"):
        return _parse_xlsx(filepath, account_id, import_id)
    return _parse_csv(filepath, account_id, import_id)


def _parse_xlsx(filepath: str, account_id: str, import_id: int) -> list[dict]:
    from openpyxl import load_workbook
    transactions = []

    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Map header → index
    header = [str(c or "").strip() for c in rows[0]]
    def col(name):
        try: return header.index(name)
        except ValueError: return None

    idx_date     = col("Transaction date")
    idx_desc     = col("Description")
    idx_merchant = col("Merchant")
    idx_card     = col("Card number")
    idx_currency = col("Currency")
    idx_amount   = col("Amount")
    idx_fcurr    = col("Foreign Currency")
    idx_famount  = col("Amount in foreign currency")
    idx_dc       = col("Debit/Credit")
    idx_status   = col("Status")

    for row in rows[1:]:
        if not row or all(v is None for v in row):
            continue

        status = str(row[idx_status] or "").strip() if idx_status is not None else ""
        if status.lower() not in ("posted", "booked", ""):
            continue

        date_val = row[idx_date] if idx_date is not None else None
        if date_val is None:
            continue

        try:
            if isinstance(date_val, datetime):
                date_str = date_val.strftime("%Y-%m-%d")
            else:
                date_str = str(date_val)[:10]
        except Exception:
            continue

        merchant = str(row[idx_merchant] or "").strip() if idx_merchant is not None else ""
        desc     = str(row[idx_desc]     or "").strip() if idx_desc     is not None else ""
        currency = str(row[idx_currency] or "CHF").strip() if idx_currency is not None else "CHF"
        dc       = str(row[idx_dc]       or "Debit").strip() if idx_dc is not None else "Debit"

        try:
            amount_raw = float(row[idx_amount]) if idx_amount is not None else 0.0
        except (TypeError, ValueError):
            continue

        # Debit = expense (negative), Credit = income/refund (positive)
        if dc.lower() == "credit":
            amount_chf = abs(amount_raw)
            is_internal = 0
        else:
            amount_chf = -abs(amount_raw)
            is_internal = 0

        # CC settlement
        raw_text = f"{desc} {merchant}".strip()
        if "your payment" in raw_text.lower() or "payment - thank you" in raw_text.lower():
            is_internal = 1

        # Foreign currency
        o_curr = str(row[idx_fcurr] or "").strip() if idx_fcurr is not None else ""
        try:
            o_amt = float(row[idx_famount]) if idx_famount is not None and row[idx_famount] else None
        except (TypeError, ValueError):
            o_amt = None

        o_curr = o_curr if o_curr and o_curr != currency else None

        # XLSX Merchant field is already clean — use directly, title-case
        merchant_clean = merchant.title() if merchant else extract_merchant_swisscard(desc, "")

        is_sub = 1 if re.search(
            r'\*SUBSCRIPTION|\*ANNUAL|HBR\*|PRIME|NETFLIX|SPOTIFY|DISNEY',
            raw_text, re.IGNORECASE
        ) else 0

        tx_id = make_tx_id("SWISSCARD", date_str, amount_chf, _merchant_key(merchant))

        transactions.append({
            "id":             tx_id,
            "date":           date_str,
            "account_id":     account_id,
            "amount":         amount_chf,
            "currency":       "CHF",
            "amount_orig":    o_amt,
            "currency_orig":  o_curr,
            "fx_rate":        None,
            "raw_text":       raw_text,
            "merchant_clean": merchant_clean,
            "tx_type":        "CARD",
            "l1":             "",
            "l2":             "",
            "is_recurring":   0,
            "is_sub":         is_sub,
            "is_internal":    is_internal,
            "import_id":      import_id,
        })

    return transactions


def _parse_csv(filepath: str, account_id: str, import_id: int) -> list[dict]:
    """Parse original Swisscard CSV format."""
    transactions = []

    with open(filepath, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)

        for row in reader:
            tx_id_raw   = (row.get("TransactionId") or "").strip()
            date_raw    = (row.get("Date") or "").strip()
            amount_raw  = (row.get("Amount") or "0").strip()
            currency    = (row.get("Currency") or "CHF").strip()
            orig_amount = (row.get("OriginalAmount") or "").strip()
            orig_curr   = (row.get("OriginalCurrency") or "").strip()
            fx_rate_raw = (row.get("Exchange Rate") or "1").strip()
            merchant    = (row.get("MerchantName") or "").strip()
            details     = (row.get("Details") or "").strip()
            state       = (row.get("StateType") or "").strip()

            if state and state.upper() != "BOOKED":
                continue
            if not tx_id_raw or not date_raw:
                continue

            try:
                date_str = datetime.fromisoformat(date_raw).strftime("%Y-%m-%d")
            except Exception:
                date_str = date_raw[:10]

            try:
                amount_chf = float(amount_raw)
            except ValueError:
                continue

            raw_text = f"{merchant} {details}".strip()
            is_internal = 0
            if amount_chf < 0 or "your payment" in raw_text.lower():
                is_internal = 1
            else:
                amount_chf = -abs(amount_chf)

            try:
                fx = float(fx_rate_raw)
            except ValueError:
                fx = 1.0

            try:
                o_amt = float(orig_amount) if orig_amount else None
            except ValueError:
                o_amt = None

            o_curr = orig_curr if orig_curr and orig_curr != currency else None
            # Use MerchantName directly for consistent dedup with XLSX
            merchant_clean = merchant.title() if merchant else extract_merchant_swisscard(details, "")

            is_sub = 1 if re.search(
                r'\*SUBSCRIPTION|\*ANNUAL|HBR\*|PRIME|NETFLIX|SPOTIFY|DISNEY',
                details, re.IGNORECASE
            ) else 0

            tx_id = make_tx_id("SWISSCARD", date_str, amount_chf, _merchant_key(merchant))

            transactions.append({
                "id":            tx_id,
                "date":          date_str,
                "account_id":    account_id,
                "amount":        amount_chf,
                "currency":      "CHF",
                "amount_orig":   o_amt,
                "currency_orig": o_curr,
                "fx_rate":       fx if fx != 1.0 else None,
                "raw_text":      raw_text,
                "merchant_clean": merchant_clean,
                "tx_type":       "CARD",
                "l1":            "",
                "l2":            "",
                "is_recurring":  0,
                "is_sub":        is_sub,
                "is_internal":   is_internal,
                "import_id":     import_id,
            })

    return transactions

