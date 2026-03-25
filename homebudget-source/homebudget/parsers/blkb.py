"""
parsers/blkb.py — BLKB Basellandschaftliche Kantonalbank XLSX parser

File format:
  - Row 0: header (Auftragsdatum, Buchungstext, Betrag Einzelzahlung (CHF),
                   Belastungsbetrag (CHF), Gutschriftsbetrag (CHF),
                   Valutadatum, Saldo (CHF))
  - No pre-table metadata
  - Account identity from filename IBAN e.g. CH38_3076_9056_9129_9200_3_...xlsx

Sign convention in source:
  Belastungsbetrag = debit (positive in file → negative in DB)
  Gutschriftsbetrag = credit (positive in file → positive in DB)
"""
import re
import os
from datetime import datetime
from openpyxl import load_workbook

from normalizer import detect_tx_type, extract_merchant_blkb
from db import make_tx_id


def extract_iban_from_filename(filename: str) -> str:
    """
    Extract IBAN from filename.
    Handles spaces, underscores, and extra suffixes like " copy":
      "CH38 3076 9056 9129 9200 3 20260321_1103_Buchungsliste copy.xlsx"
      "CH38_3076_9056_9129_9200_3_20260321_1103_Buchungsliste.xlsx"
    → "CH3830769056912992003"
    """
    base = os.path.basename(filename)
    # Find CH + sequence of digits/spaces/underscores
    m = re.search(r'(CH\d{2}[\d\s_]+)', base, re.IGNORECASE)
    if m:
        iban = re.sub(r'[\s_]', '', m.group(1)).upper()
        # Swiss IBAN is 21 chars — trim trailing digits if too long
        if len(iban) > 21:
            iban = iban[:21]
        return iban
    return ""


def is_blkb_file(filepath: str) -> bool:
    """Detect if file is a BLKB export."""
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        headers = [str(c.value or "").strip() for c in next(ws.iter_rows(max_row=1))]
        wb.close()
        return "Buchungstext" in headers and "Belastungsbetrag (CHF)" in headers
    except Exception:
        return False


def parse(filepath: str, account_id: str, import_id: int) -> list[dict]:
    """
    Parse BLKB XLSX file.
    Returns list of transaction dicts ready for insert_transaction().
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # Header row
    header = [str(h or "").strip() for h in rows[0]]

    def col(name):
        try:
            return header.index(name)
        except ValueError:
            return None

    idx_date    = col("Auftragsdatum")
    idx_text    = col("Buchungstext")
    idx_debit   = col("Belastungsbetrag (CHF)")
    idx_credit  = col("Gutschriftsbetrag (CHF)")

    if None in (idx_date, idx_text, idx_debit, idx_credit):
        raise ValueError(f"BLKB parser: unexpected header format: {header}")

    transactions = []

    for row in rows[1:]:
        raw_date  = row[idx_date]
        raw_text  = str(row[idx_text] or "").strip()
        debit     = row[idx_debit]
        credit    = row[idx_credit]

        if not raw_text or raw_date is None:
            continue

        # Normalise date
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            try:
                date_str = datetime.strptime(str(raw_date), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except Exception:
                date_str = str(raw_date)[:10]

        # Amount — debit negative, credit positive
        if debit and float(debit) > 0:
            amount = -abs(float(debit))
        elif credit and float(credit) > 0:
            amount = abs(float(credit))
        else:
            continue

        # Detect type + extract merchant
        tx_type  = detect_tx_type(raw_text)
        merchant = extract_merchant_blkb(raw_text, tx_type)

        # Internal transfer detection
        is_internal = 1 if tx_type == "INTERNAL_TRANSFER" else 0

        tx_id = make_tx_id("BLKB", date_str, amount, raw_text)

        transactions.append({
            "id":           tx_id,
            "date":         date_str,
            "account_id":   account_id,
            "amount":       amount,
            "currency":     "CHF",
            "amount_orig":  None,
            "currency_orig": None,
            "fx_rate":      None,
            "raw_text":     raw_text,
            "merchant_clean": merchant,
            "tx_type":      tx_type,
            "l1":           "",
            "l2":           "",
            "is_recurring": 0,
            "is_sub":       0,
            "is_internal":  is_internal,
            "import_id":    import_id,
        })

    return transactions
